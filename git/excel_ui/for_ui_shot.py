import shotgun_api3
from PyQt5 import QtWidgets
# from PyQt5.QtWidgets import QApplication, QWidget
from handler_f import ShotgunAction
import sys
import openpyxl


class CreateXlsDialog(QtWidgets.QDialog):
    """
    A class that creates a QDialog window to prompt the user for a file path to save an Excel file.

    Attributes:
    okBtn (QPushButton): The button to accept the dialog.
    pathBrowseBtn (QPushButton): The button to open a file browser and choose a file path.
    pathEdit (QLineEdit): The text box to display the chosen file path.
    pathLabel (QLabel): The label to display the text "Path:".
    path (str): The chosen file path.

    """

    def __init__(self, parent=None):
        super(CreateXlsDialog, self).__init__(parent)
        self.okBtn = None
        self.pathBrowseBtn = None
        self.pathEdit = None
        self.pathLabel = None
        self.path = None
        # self.action = ShotgunAction(sys.argv[1])
        sa = ShotgunAction(sys.argv[1])
        self.selected_ids = sa.data()
        self.initUI()

    def initUI(self):
        """
        A method to initialize the UI of the QDialog window.

        """
        self.pathLabel = QtWidgets.QLabel("Path:")
        self.pathEdit = QtWidgets.QLineEdit()
        self.pathBrowseBtn = QtWidgets.QPushButton(">>")
        self.pathBrowseBtn.clicked.connect(self.browsePath)

        self.okBtn = QtWidgets.QPushButton("OK")
        self.okBtn.clicked.connect(self.create_xls)

        layout = QtWidgets.QHBoxLayout()
        layout.addWidget(self.pathLabel)
        layout.addWidget(self.pathEdit)
        layout.addWidget(self.pathBrowseBtn)

        buttonLayout = QtWidgets.QHBoxLayout()
        buttonLayout.addStretch()
        buttonLayout.addWidget(self.okBtn)

        mainLayout = QtWidgets.QVBoxLayout()
        mainLayout.addLayout(layout)
        mainLayout.addLayout(buttonLayout)

        self.setLayout(mainLayout)
        self.setWindowTitle("Create XLS")

    def browsePath(self):
        """
        A method to open a file browser and choose a file path.

        """
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        file_name, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save file", "", "Excel Files (*.xlsx)",
                                                             options=options)
        if file_name:
            self.pathEdit.setText(file_name)
            self.path = file_name

    def getPath(self):
        """
        A method to get the chosen file path.

        Returns:
        str: The chosen file path.

       """
        return self.path

    def accept(self):
        """
        A method to accept the dialog and close it.

        """
        super().accept()

    # def data(self):
    #     pass

    def create_xls(self):
        """
        A function to create an Excel file with the given entity name and status.

        Args:
        entity_name (str): The name of the entity to be included in the Excel file.
        entity_status (str): The status of the entity to be included in the Excel file.

        """

        SERVER_PATH = 'https://rndtest.shotgrid.autodesk.com'
        SCRIPT_NAME = 'script_wongyu'
        SCRIPT_KEY = 'kmjozapdzo7vyk~bqvlLsxgsn'

        sg = shotgun_api3.Shotgun(SERVER_PATH, SCRIPT_NAME, SCRIPT_KEY)
        # ws = wb.active
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Version Name")
        ws.cell(row=1, column=2, value="Status")
        row_num = 2
        # versions = sg.find("Version", [], ["code", "sg_status_list"])
        for selected_id in self.selected_ids:
            versions = sg.find_one("Shot", [["id", "is", int(selected_id)]], ["id", "code", "sg_status_list"])
            ws.cell(row=row_num, column=1, value=versions["code"])
            # ws.cell(row=row_num, column=1, value=versions.get("code"))
            ws.cell(row=row_num, column=2, value=versions["sg_status_list"])
            row_num += 1
            print(versions["code"])

        # app = QtWidgets.QApplication(sys.argv)
        # dlg = CreateXlsDialog()
        # if dlg.exec_():
        # wb = openpyxl.Workbook()
        # ws = wb.active

        # Write headers to worksheet
        # ws.cell(row=1, column=1, value="Version Name")
        # ws.cell(row=1, column=2, value="Status")
        #
        # # Write version data to worksheet
        #
        #
        # # ws.append([self.action.code])
        # row_num = 2
        # ws.cell(row=row_num, column=1, value=versions["code"])
        # # ws.cell(row=row_num, column=1, value=versions.get("code"))
        # ws.cell(row=row_num, column=2, value=versions["sg_status_list"])
        # row_num += 1
        #
        # for version in versions:
        #     print("123456789", versions)
        #     print("123", version)
        #     # ws.cell(row=row_num, column=1, value=version["code"])
        #     ws.cell(row=row_num, column=1, value=version["code"])
        #     # ws.cell(row=row_num, column=2, value=version["sg_status_list"])
        #     row_num += 1
        try:
            wb.save(self.path + '.xlsx')
            print("saved")
        except Exception as e:
            print(f"error: {str(e)}")
        # sys.exit(app.exec_())


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    cd = CreateXlsDialog()
    cd.show()
    sys.exit(app.exec_())


    # sg_status_list = "example_status"
    # create_xls(name, sg_status_list)
