from PyQt5 import QtWidgets
import sys
import openpyxl
import shotgun_api3
from shotgun_api3 import Shotgun
import xlsxwriter

SERVER_PATH = 'https://rndtest.shotgrid.autodesk.com'
SCRIPT_NAME = 'script_wongyu'
SCRIPT_KEY = 'stunj5lyzpkQyeapzrbdejd-b'

class CreateXlsDialog(QtWidgets.QDialog):
    """
    A class that creates a QDialog window to prompt the user for a file path to save an Excel file.

    Attributes:
    okBtn (QPushButton): The button to accept the dialog.
    pathBrowseBtn (QPushButton): The button to open a file browser and choose a file path.
    pathEdit (QLineEdit): The text box to display the chosen file path.
    pathLabel (QLabel): The label to display the text "Path:".
    path (str): The chosen file path.

    """

    def __init__(self, parent=None):
        super(CreateXlsDialog, self).__init__(parent)
        self.okBtn = None
        self.pathBrowseBtn = None
        self.pathEdit = None
        self.pathLabel = None
        self.path = None
        self.initUI()

    def initUI(self):
        """
        A method to initialize the UI of the QDialog window.

        """
        self.pathLabel = QtWidgets.QLabel("Path:")
        self.pathEdit = QtWidgets.QLineEdit()
        self.pathBrowseBtn = QtWidgets.QPushButton(">>")
        self.pathBrowseBtn.clicked.connect(self.browsePath)

        self.okBtn = QtWidgets.QPushButton("OK")
        self.okBtn.clicked.connect(self.accept)

        layout = QtWidgets.QHBoxLayout()
        layout.addWidget(self.pathLabel)
        layout.addWidget(self.pathEdit)
        layout.addWidget(self.pathBrowseBtn)

        buttonLayout = QtWidgets.QHBoxLayout()
        buttonLayout.addStretch()
        buttonLayout.addWidget(self.okBtn)

        mainLayout = QtWidgets.QVBoxLayout()
        mainLayout.addLayout(layout)
        mainLayout.addLayout(buttonLayout)

        self.setLayout(mainLayout)
        self.setWindowTitle("Create XLS")

    def browsePath(self):
        """
        A method to open a file browser and choose a file path.

        """
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        file_name, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save file", "", "Excel Files (*.xlsx)",
                                                             options=options)
        if file_name:
            self.pathEdit.setText(file_name)
            self.path = file_name

    def getPath(self):
        """
        A method to get the chosen file path.

        Returns:
        str: The chosen file path.

       """
        return self.path

    def accept(self):
        """
        A method to accept the dialog and close it.

        """
        super().accept()


def create_xls(entity_name, entity_status, PROJECT_ID=None):
    """
    A function to create an Excel file with the given entity name and status.

    Args:
    entity_name (str): The name of the entity to be included in the Excel file.
    entity_status (str): The status of the entity to be included in the Excel file.

    """
    sg = Shotgun(SERVER_PATH, SCRIPT_NAME, SCRIPT_KEY)
    # ver = sg.find("Version", filters=[["sg_status_list", "is", "ip"]], fields=["code", "sg_status_list"])
    # filters = [
    #     ["project", "is", {"type": "Project", "id": PROJECT_ID}],
    #     ["entity", "type_is", "Asset"],
    #     ["entity", "name_is", entity_name],
    #     ["sg_status_list", "is", entity_status]
    # ]
    # fields = ["code", "sg_status_list"]
    ver = sg.find("Version", filters=[
            ["project", "is", {"type": "Project", "id": PROJECT_ID}],
            ["code", "is", entity_name],
            ["sg_status_list", "is", entity_status]
        ],
        fields=["code", "sg_status_list"]
    )
    app = QtWidgets.QApplication(sys.argv)
    dlg = CreateXlsDialog()
    if dlg.exec_():
        path = dlg.getPath()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Code", "Status"])
        for asset in asset_data:
            ws.append([asset["code"], asset["sg_status_list"]])
        # ws.title = "Entity Info"
        # ws["A1"] = "Name"
        # ws["B1"] = "Status"
        # ws["A2"] = entity_name
        # ws["B2"] = entity_status
        try:
            wb.save(path + '.xlsx')
            print("saved")
        except Exception as e:
            print(f"error: {str(e)}")
    sys.exit(app.exec_())


if __name__ == '__main__':
    create_xls()
