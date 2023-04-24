from PySide2 import QtWidgets, QtGui, QtCore
import sys
import openpyxl
from shotgun_api3 import Shotgun
import sgtk


class CreateXlsDialog(QtWidgets.QDialog):
    """
    A class that creates a QDialog window to prompt the user for a file path to save an Excel file.

    Attributes:
    okBtn (QPushButton): The button to accept the dialog.
    pathBrowseBtn (QPushButton): The button to open a file browser and choose a file path.
    pathEdit (QLineEdit): The text box to display the chosen file path.
    pathLabel (QLabel): The label to display the text "Path:".
    path (str): The chosen file path.

    """

    def __init__(self, parent=None):
        super(CreateXlsDialog, self).__init__(parent)
        self.okBtn = None
        self.pathBrowseBtn = None
        self.pathEdit = None
        self.pathLabel = None
        self.path = None
        self.initUI()

    def initUI(self):
        """
        A method to initialize the UI of the QDialog window.

        """
        self.pathLabel = QtWidgets.QLabel("Path:")
        self.pathEdit = QtWidgets.QLineEdit()
        self.pathBrowseBtn = QtWidgets.QPushButton(">>")
        self.pathBrowseBtn.clicked.connect(self.browsePath)

        self.okBtn = QtWidgets.QPushButton("OK")
        self.okBtn.clicked.connect(self.accept)

        layout = QtWidgets.QHBoxLayout()
        layout.addWidget(self.pathLabel)
        layout.addWidget(self.pathEdit)
        layout.addWidget(self.pathBrowseBtn)

        buttonLayout = QtWidgets.QHBoxLayout()
        buttonLayout.addStretch()
        buttonLayout.addWidget(self.okBtn)

        mainLayout = QtWidgets.QVBoxLayout()
        mainLayout.addLayout(layout)
        mainLayout.addLayout(buttonLayout)

        self.setLayout(mainLayout)
        self.setWindowTitle("Create XLS")

    def browsePath(self):
        """
        A method to open a file browser and choose a file path.

        """
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        file_name, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save file", "", "Excel Files (*.xlsx)",
                                                             options=options)
        if file_name:
            self.pathEdit.setText(file_name)
            self.path = file_name

    def getPath(self):
        """
        A method to get the chosen file path.

        Returns:
        str: The chosen file path.

       """
        return self.path


def create_xls(entity_name, entity_status):
    """
    A function to create an Excel file with the given entity name and status.

    Args:
    entity_name (str): The name of the entity to be included in the Excel file.
    entity_status (str): The status of the entity to be included in the Excel file.

    """
    dlg = CreateXlsDialog()
    if dlg.exec_():
        path = dlg.getPath()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Entity Info"
        ws["A1"] = "Entity Name"
        ws["B1"] = "Status"
        ws["A2"] = entity_name
        ws["B2"] = entity_status
        wb.save(path + '.xlsx')

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    create_xls("Entity 1", "Active")
    sys.exit(app.exec_())


# except IndexError:
# print("Missing Get arguments")
# import openpyxl
# from PySide2 import QtWidgets
# from PySide2.QtCore import QUrl
# from PySide2.QtGui import QDesktopServices
# from urllib.parse import urlparse
# import os
# import shotgun_api3
# 
# class CreateXlsDialog(QtWidgets.QDialog):
#     def __init__(self, parent=None):
#         super(CreateXlsDialog, self).__init__(parent)
#         self.path = None
#         self.initUI()
# 
#     def initUI(self):
#         self.pathLabel = QtWidgets.QLabel("Path:")
#         self.pathEdit = QtWidgets.QLineEdit()
#         self.pathBrowseBtn = QtWidgets.QPushButton(">>")
#         self.pathBrowseBtn.clicked.connect(self.browsePath)
# 
#         self.okBtn = QtWidgets.QPushButton("OK")
#         self.okBtn.clicked.connect(self.accept)
# 
#         layout = QtWidgets.QHBoxLayout()
#         layout.addWidget(self.pathLabel)
#         layout.addWidget(self.pathEdit)
#         layout.addWidget(self.pathBrowseBtn)
# 
#         buttonLayout = QtWidgets.QHBoxLayout()
#         buttonLayout.addStretch()
#         buttonLayout.addWidget(self.okBtn)
# 
#         mainLayout = QtWidgets.QVBoxLayout()
#         mainLayout.addLayout(layout)
#         mainLayout.addLayout(buttonLayout)
# 
#         self.setLayout(mainLayout)
#         self.setWindowTitle("Create XLS")
# 
#     def browsePath(self):
#         options = QtWidgets.QFileDialog.Options()
#         options |= QtWidgets.QFileDialog.DontUseNativeDialog
#         file_name, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save file", "", "Excel Files (*.xlsx)",
#                                                              options=options)
#         if file_name:
#             self.pathEdit.setText(file_name)
#             self.path = file_name
# 
#     def getPath(self):
#         return self.path
# 
# 
# def create_xls(entity_name, entity_status):
#     dlg = CreateXlsDialog()
#     if dlg.exec_():
#         path = dlg.getPath()
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Entity Info"
#         ws["A1"] = "Entity Name"
#         ws["B1"] = "Status"
#         ws["A2"] = entity_name
#         ws["B2"] = entity_status
#         wb.save(path)
# 
# #         
# if __name__ == '__main__':
#     app = QtWidgets.QApplication(sys.argv)
#     main_window = MainWindow()
#     main_window.show()
#     create_xls(main_window.entity_name, main_window.entity_status)
#     sys.exit(app.exec_())
