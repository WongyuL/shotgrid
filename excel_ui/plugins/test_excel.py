from PySide2 import QtWidgets
import openpyxl
from PySide2.QtCore import QUrl
from PySide2.QtGui import QDesktopServices
from urllib.parse import urlparse
import shotgun_api3
from openpyxl.workbook import Workbook
import sys
import pprint
import os
from openpyxl import Workbook
import sgtk

try:
    from urlparse import parse_qs
except ImportError:
    from urllib.parse import parse_qs


class CreateXlsDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(CreateXlsDialog, self).__init__(parent)
        self.okBtn = None
        self.pathBrowseBtn = None
        self.pathEdit = None
        self.pathLabel = None
        self.path = None
        self.initUI()

    def initUI(self):
        self.pathLabel = QtWidgets.QLabel("Path:")
        self.pathEdit = QtWidgets.QLineEdit()
        self.pathBrowseBtn = QtWidgets.QPushButton(">>")
        self.pathBrowseBtn.clicked.connect(self.browsePath)

        self.okBtn = QtWidgets.QPushButton("OK")
        self.okBtn.clicked.connect(self.accept)

        layout = QtWidgets.QHBoxLayout()
        layout.addWidget(self.pathLabel)
        layout.addWidget(self.pathEdit)
        layout.addWidget(self.pathBrowseBtn)

        buttonLayout = QtWidgets.QHBoxLayout()
        buttonLayout.addStretch()
        buttonLayout.addWidget(self.okBtn)

        mainLayout = QtWidgets.QVBoxLayout()
        mainLayout.addLayout(layout)
        mainLayout.addLayout(buttonLayout)

        self.setLayout(mainLayout)
        self.setWindowTitle("Create XLS")

    def browsePath(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        file_name, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save file", "", "Excel Files (*.xlsx)",
                                                             options=options)
        if file_name:
            self.pathEdit.setText(file_name)
            self.path = file_name

    def getPath(self):
        return self.path


# Get the entity type and id from the URL passed in as an argument
def create_xls(url):
    url = sys.argv[1]
    url_components = urlparse(url)
    entity_type = None
    if len(url_components.path.split("/")) > 1:
        entity_type = url_components.path.split("/")[-2]

    # If there is a querystring, parse it
    if url_components.query:
        query_params = parse_qs(url_components.query)
        entity_id = query_params.get("entity_id")[0]
        entity_name = query_params.get("entity_name")[0]
    else:
        entity_id = None
        entity_name = None

    # Connect to ShotGrid
    sg = shotgun_api3.Shotgun(os.environ.get("https://rndtest.shotgrid.autodesk.com"),
                              script_name=os.environ.get("script_wongyu"),
                              api_key=os.environ.get("mWum6vnnrxfn%wtaoutismjiw"))

    # Get the entity's name and status
    if entity_id and entity_type:
        entity = sg.find_one(entity_type, [["id", "is", entity_id]], ["code", "sg_status_list"])
        entity_name = entity["code"]
        entity_status = entity["sg_status_list"]

    # Create Excel file and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = entity_type

    # Write data to Excel worksheet
    if entity_type:
        data = sg.find(entity_type, [], ["code", "sg_status_list"])
        for i, record in enumerate(data):
            ws.cell(row=i + 1, column=1, value=record["code"])
            ws.cell(row=i + 1, column=2, value=record["sg_status_list"])

    wb.save(f"{entity_type}.xlsx")


def launch_ami():
    # Construct the ShotGrid URL with the appropriate parameters
    sg_url = 'ShotGrid://<entity_type>/<entity_id>?fields=<field1>,<field2>'
    sgtk.platform.current_engine().execute_command('create_xls_file', sg_url)

button = QPushButton('Create XLS', parent=main_window)
button.clicked.connect(launch_ami)


# def create_xls(entity_name=None):
#     # # sg_url = os.environ.get("https://rndtest.shotgrid.autodesk.com/page/7434")
#     # # sg_script = os.environ.get("script_wongyu")
#     # # sg_key = os.environ.get("mWum6vnnrxfn%wtaoutismjiw")
#     # sg = shotgun_api3.Shotgun("https://rndtest.shotgrid.autodesk.com",
#     #                           script_name="script_wongyu",
#     #                           api_key="xtiofjkbgu_xDhbwhwwvgpq5p")
#
#     # Get the entity type and id from the URL passed in as an argument
#     url = sys.argv[1]
#     url_components = urlparse(url)
#     entity_type = None
#     if len(url_components.path.split("/")) > 1:
#         entity_type = url_components.path.split("/")[-2]
#     sg = shotgun_api3.Shotgun("https://rndtest.shotgrid.autodesk.com", script_name="script_wongyu", api_key="xtiofjkbgu_xDhbwhwwvgpq5p")
#     data = sg.find(entity_type, [], ["code", "sg_status_list"])
#
#     #Create Excel file and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = entity_type
#
#     #Write data to Excel worksheet
#     for i, record in enumerate(data):
#         ws.cell(row=i+1, column=1, value=record["code"])
#         ws.cell(row=i+1, column=2, value=record["sg_status_list"])
#
#     wb.save(f"{entity_type}.xlsx")
#
#     # Get the entity's name and status
#     # entity = sg.find_one(entity_type, [["id", "is", entity_id]], ["code", "sg_status_list"])
#     # entity_name = entity["code"]
#     # entity_status = entity["sg_status_list"]
#
#     # Create the Excel file and write the entity's name and status to it
#     dlg = CreateXlsDialog()
#     if dlg.exec_():
#         path = dlg.getPath()
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Entity Info"
#         ws["A1"] = "Entity Name"
#         ws["B1"] = "Status"
#         ws["A2"] = entity_name
#         ws["B2"] = entity_status
#         wb.save(path)


if __name__ == '__main__':
    app = QtWidgets.QApplication()
    create_xls()
    sys.exit(app.exec_())
