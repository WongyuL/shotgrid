import sys
import openpyxl
from PySide2 import QtWidgets
from PySide2.QtCore import QUrl
from PySide2.QtGui import QDesktopServices
from urllib.parse import urlparse
import os
import shotgun_api3
from mainwindow import MainWindow

from PySide2 import QtWidgets, QtCore
import sys


class CreateXlsDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(CreateXlsDialog, self).__init__(parent)
        self.path = None
        self.initUI()

    def initUI(self):
        self.pathLabel = QtWidgets.QLabel("Path:")
        self.pathEdit = QtWidgets.QLineEdit()
        self.pathBrowseBtn = QtWidgets.QPushButton(">>")
        self.pathBrowseBtn.clicked.connect(self.browsePath)

        self.okBtn = QtWidgets.QPushButton("OK")
        self.okBtn.clicked.connect(self.accept)

        layout = QtWidgets.QHBoxLayout()
        layout.addWidget(self.pathLabel)
        layout.addWidget(self.pathEdit)
        layout.addWidget(self.pathBrowseBtn)

        buttonLayout = QtWidgets.QHBoxLayout()
        buttonLayout.addStretch()
        buttonLayout.addWidget(self.okBtn)

        mainLayout = QtWidgets.QVBoxLayout()
        mainLayout.addLayout(layout)
        mainLayout.addLayout(buttonLayout)

        self.setLayout(mainLayout)
        self.setWindowTitle("Create XLS")

    def browsePath(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        file_name, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save file", "", "Excel Files (*.xlsx)",
                                                             options=options)
        if file_name:
            self.pathEdit.setText(file_name)
            self.path = file_name

    def getPath(self):
        return self.path


def create_xls(entity_name, entity_status, path):
    dlg = CreateXlsDialog()
    if dlg.exec_():
        path = dlg.getPath()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Entity Info"
        ws["A1"] = "Entity Name"
        ws["B1"] = "Status"
        ws["A2"] = entity_name
        ws["B2"] = entity_status
        wb.save(path)


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    create_xls()
    sys.exit(app.exec_())
